# Модуль генерирует файл с Заданием № 06 на тему "Регенерация цифрового 
# сигнала при передаче на большие расстояния"
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
from random import random
from pprint import pprint as pp
import pytils.translit
import re
import numpy as np
import error_rate as er
import linear_codes as lc

def has_numbers(s):
    return re.search('\d', s)

def generator(group, student, task_code):
    # Общее число участков, N
    # Число участков с НРП, M
    while True:
        n = round( random() * (n_max - n_min) + n_min )
        m = round( random() * (m_max - m_min) + m_min )
        if n % m == 0 and m < n // 2:
            break

    print(f'Общее число участков: {n}')
    print(f'Число участков с НРП: {m}')

    while True:
        # Сигнал-шум на входе первого МШУ, q1, дБ
        q1 = round(random() * (q1_max - q1_min) + q1_min, 2)

        # Коэффициент шума, F, дБ
        f = round(random() * (f_max - f_min) + f_min, 2)

        q2_target = er.dB2pow(q1) / m / er.dB2pow(f)
        p_orp_1 = 0.5 * np.exp(- 0.5 * q2_target)
        p_target_approx, _ = lc.probability_bsc_more(0, m + 1, p_orp_1)
        if p_target_approx > 1.e-9 and p_target_approx < 0.1: 
            # Исключаем вероятность выше 0.1 
            # из-за неработоспособности такой системы передачи информации
            break

    print(f'Сигнал-шум на входе первого МШУ: {q1}')
    print(f'Коэффициент шума: {f}')
    print(f'P ош. = {p_target_approx}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'

    hf = Font(name = 'Calibri', bold = True)

    ws.append(['Общее число участков, N:'])
    ws.append([n])
    ws.append(['Число участков с НРП, M:'])
    ws.append([m])
    ws.append(['Сигнал-шум на входе первого МШУ, q1, дБ:'])
    ws.append([q1])
    ws.append(['Коэффициент шума, F, дБ:'])
    ws.append([f])

    wsC = wb.create_sheet('Check')
    wsC.append(['Введите ответы:'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf
    wsC.append(['Вероятность битовой ошибки на выходе системы передачи \
информации, приближенная формула, P ош. итоговая:'])
    wsC.append([0.])
    wsC.append(['Вероятность битовой ошибки на выходе системы передачи \
информации, точная формула, P ош. итоговая:'])
    wsC.append([0.])
    wsC.append(['Требуемое отношение сигнал-шум на входе первого МШУ, \
при условии, что все блоки - НРП, q1, дБ:'])
    wsC.append([0.])
    wsC.append(['Требуемое отношение сигнал-шум на входе первого МШУ, \
при условии, что все блоки - ОРП, q1, дБ:'])
    wsC.append([0.])
    wsC.append(['Внимание! Ответы давать с точностью не хуже 1%'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf

    wb.save(f'{student}_{task_code}_{group}.xlsx')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    fn = 'list_magister_titpi_2020.txt'
    task_code = '06'

    # Ограничения на общее число участков, N
    n_min = 6
    n_max = 81
    # Ограничения на число участков с НРП, M
    m_min = 2
    m_max = 81

    # Ограничения на сигнал-шум на входе первого МШУ, q1, дБ
    q1_min = 15
    q1_max = 35

    # Ограничения на коэффициент шума, F, дБ
    f_min = 1.0
    f_max = 3.0

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if s:
            s_translit = pytils.translit.translify(s)
            print(s_translit)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                generator(group, student, task_code)
