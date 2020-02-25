# Модуль для проверки ответов к Заданию № 01
from openpyxl import load_workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
import operator
from functools import reduce
from pprint import pprint as pp
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

# Проверяет состоит ли вектор только из 0 и 1
def is_bits_vector(v):
    return all(map(lambda x: (x == 0) or (x == 1), v))

# Проверяет состоит ли вектор только из чисел типа int
def is_int_vector(v):
    return all(map(lambda x: isinstance(x, int), v))

def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return

    ws = wb['Main']
    G = []
    parameters = []
    for row in ws.iter_rows(min_row = 1, max_col = max_n, max_row = 16 + max_k, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        g_Ok = is_bits_vector(row)
        n = len(row)
        if g_Ok and n >= min_n and n <= max_n:
            G.append(row) # Читаем порождающую матрицу G кода
        if n == 1 and isinstance(row[0], int):
            parameters.append(row[0])

    print('Порождающая матрица кода')
    pp(G)

    k, n, ok = lc.check_matrix(G)
    r = n - k

    assert((len(parameters) == 6) and ok)

    n_, k_, r_, d_, qo_, qi_ = parameters

    Wsp = lc.gen_spectrum(G)
    d = lc.spectrum_to_code_distance(Wsp)
    qo = d - 1
    qi = (d - 1) // 2

    print('Правильные ответы:')
    pp(f'n = {n}')
    pp(f'k = {k}')
    pp(f'r = {r}')
    pp(f'd = {d}')
    pp(f'qo = {qo}')
    pp(f'qi = {qi}')

    print('Введенные ответы:')
    pp(f'n = {n_}')
    pp(f'k = {k_}')
    pp(f'r = {r_}')
    pp(f'd = {d_}')
    pp(f'qo = {qo_}')
    pp(f'qi = {qi_}')

    assert(n_ == n)
    assert(k_ == k)
    assert(r_ == r)
    assert(d_ == d)
    assert(qo_ == qo)
    assert(qi_ == qi)


    wsSp = wb['CodeSpectrum']
    tmp_ = []
    for row in wsSp.iter_rows(min_row = 1, max_col = max_n + 1, max_row = 5, values_only = True):
        row = list(filter(None.__ne__, row))
        g_Ok = is_int_vector(row)
        n = len(row)
        if g_Ok and n <= max_n + 1:
            tmp_.append(row)

    assert(len(tmp_) == 2)

    spC_ = dict((k, v) for k, v in zip(tmp_[0], tmp_[1]) if v > 0)

    print('Правильный спектр кода')
    pp(Wsp)

    print('Введенный спектр кода')
    pp(spC_)

    assert(spC_ == Wsp)

    pp('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '01'
    fn = 'list_magister_titpi_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 6
    max_n = 15
    min_k = 3
    max_k = 5
    min_r = 2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)