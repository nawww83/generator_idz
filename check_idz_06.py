# Модуль для проверки ответов к Заданию № 06
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
import error_rate as er
import linear_codes as lc
from pprint import pprint as pp
import numpy as np
import scipy.optimize as opt
from random import random

mjr = sys.version_info.major
mnr = sys.version_info.minor
if (mjr == 3 and mnr < 7) or mjr < 3:
    print('Требуется Python версии 3.7 и выше!')
    exit()

student = 'IvanovAA'
task_code = '06'
group = '1B6'

# Ограничения на общее число участков, N
n_min = 6
n_max = 81
# Ограничения на число участков с НРП, M
m_min = 2
m_max = 81

# Ограничения на сигнал-шум на входе первого МШУ, q1, дБ
q1_min = 15
q1_max = 35

# Ограничения на коэффициент шума, F, дБ
f_min = 1.0
f_max = 3.0

hf = Font(name = 'Calibri', bold = True)

fname = f'{student}_{task_code}_{group}.xlsx'

wb = load_workbook(fname)

ws = wb['Main']
params = []
for row in ws.iter_rows(min_row = 1, max_col = 1, max_row = 9, values_only = True):
    row = list(filter(None.__ne__, row)) # Убирает ненужные None
    n_row = len(row)
    if n_row == 1 and (isinstance(row[0], int) or isinstance(row[0], float)):
        params.append(row[0])

n_, m_, q1_, f_ = params
assert(n_ >= n_min)
assert(n_ <= n_max)
assert(m_ >= m_min)
assert(m_ <= m_max)
assert(q1_ <= q1_max)
assert(q1_ >= q1_min)
assert(f_ <= f_max)
assert(f_ >= f_min)

print(f'Общее число участков, N: {n_}')
print(f'Число участков с НРП, M: {m_}')
print(f'Сигнал-шум на входе первого МШУ, q1, дБ: {q1_}')
print(f'Коэффициент шума, F, дБ: {f_}')

wsC = wb['Check']
params = []
for row in wsC.iter_rows(min_row = 1, max_col = 1, max_row = 10, values_only = True):
    row = list(filter(None.__ne__, row)) # Убирает ненужные None
    n_row = len(row)
    if n_row == 1 and isinstance(row[0], float) or isinstance(row[0], int):
        params.append(row[0])

p_approx_, p_exact_, q1_nrp_, q1_orp_ = params

assert(p_approx_ >= 0.)
assert(p_approx_ <= 1.)
assert(p_exact_ >= 0.)
assert(p_exact_ <= 1.)
assert(q1_nrp_ >= q1_orp_)

print(f'Введенные ответы:')
print(f'Вероятность битовой ошибки на выходе системы передачи информации')
print(f' - приближенная формула: {p_approx_}')
print(f' - точная формула: {p_exact_}')
print(f'Требуемое отношение сигнал-шум на входе первого МШУ, если все блоки:')
print(f' - НРП: {q1_nrp_}, дБ')
print(f' - ОРП: {q1_orp_}, дБ')

# Решение
q2_target = er.dB2pow(q1_) / m_ / er.dB2pow(f_)
p_orp_1 = 0.5 * np.exp(- 0.5 * q2_target)

p_target_approx, _ = lc.probability_bsc_more(0, n_ // m_, p_orp_1)
p_target_exact = 0.
for i in range(1, m_ + 2, 2):
    p_target_exact += lc.probability_bsc(i, n_ // m_, p_orp_1)

q2_nrp = - 2. * n_ * er.dB2pow(f_) * np.log(2 * p_target_exact)
p_tmp = 1. - np.power(1. - p_target_exact, 1. / (n_ + 1.))
q2_orp = - 2. * er.dB2pow(f_) * np.log(2 * p_tmp) 

print(f'Правильные ответы:')
print(f'Вероятность битовой ошибки на выходе системы передачи информации')
print(f' - приближенная формула: {p_target_approx}')
rel_error = np.abs(p_approx_ - p_target_approx) / p_target_approx
assert(rel_error < 0.01)

print(f' - точная формула: {p_target_exact}')
rel_error = np.abs(p_exact_ - p_target_exact) / p_target_exact
assert(rel_error < 0.01)

print(f'Требуемое отношение сигнал-шум на входе первого МШУ, если все блоки:')
print(f' - НРП: {round(er.pow2dB(q2_nrp), 2)}, дБ')
rel_error = np.abs((q1_nrp_ - er.pow2dB(q2_nrp)) / er.pow2dB(q2_nrp))
assert(rel_error < 0.01)

print(f' - ОРП: {round(er.pow2dB(q2_orp), 2)}, дБ')
rel_error = np.abs((q1_orp_ - er.pow2dB(q2_orp)) / er.pow2dB(q2_orp))
assert(rel_error < 0.01)


print('All Ok')

