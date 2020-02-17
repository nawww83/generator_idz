# Модуль генерирует файл с Заданием № 07 на тему "Коды Хаффмана и Шеннона-Фано"
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
from random import random
from random import randint
from pprint import pprint as pp
import pytils.translit
import re
import numpy as np
import error_rate as er
import linear_codes as lc

def has_numbers(s):
    return re.search('\d', s)

def in_range_incl(x, a, b):
    return (x >= a) and (x <= b)

def generator(group, student, task_code):
    # Вероятности троичного символа, p1, p2, p3
    p1 = p2 = p3 = 0.
    n_of_digits = int(1.5 * np.round(np.log10(1. / min_p) + 0.5))
    while not in_range_incl(p1, min_p, max_p) or \
            not in_range_incl(p2, min_p, max_p) or \
            not in_range_incl(p3, min_p, max_p):
        p1 = random()
        p2 = random()
        p3 = random()
        norma = p1 + p2 + p3
        p1 /= norma
        p2 /= norma
        p3 /= norma
        p1 = np.round(p1, n_of_digits)
        p2 = np.round(p2, n_of_digits)
        p3 = np.round(p3, n_of_digits)
        epsilon = np.power(10., -n_of_digits)
        norma = p1 + p2 + p3
        l = [p1, p2, p3]
        i = l.index(max(l))
        j = l.index(min(l))
        norma = sum(l)
        while norma != 1.:
            if norma > 1.:
                l[i] -= epsilon
            if norma < 1.:
                l[j] += epsilon
            norma = round(sum(l), n_of_digits + 1)
        p1, p2, p3 = l
        p1 = np.round(p1, n_of_digits)
        p2 = np.round(p2, n_of_digits)
        p3 = np.round(p3, n_of_digits)

    al = {0: p1, 1: p2, 2: p3}
    print(f'Алфавит: {al}, норма: {p1 + p2 + p3}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'

    hf = Font(name = 'Calibri', bold = True)

    ws.append(['Алфавит, X:'])
    ws.append(list(al.keys()))
    ws.append(list(al.values()))

    wsC = wb.create_sheet('Check')
    wsC.append(['Введите ответы:'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf
    wsC.append(['Код Хаффмана для символа X:'])
    wsC.append(['Символ', 'Код'])
    for k, v in al.items():
        len_code_limit = int(np.log2(1. / v) + 0.5)
        code = lc.get_rand_bits(len_code_limit)
        wsC.append([k] + code)
    wsC.append(['Код Хаффмана для символа XX:'])
    wsC.append(['Символ', 'Код'])
    for k1, v1 in al.items():
        for k2, v2 in al.items():
            len_code_limit = int(np.log2(1. / v1 / v2) + 0.5)
            code = lc.get_rand_bits(len_code_limit)
            wsC.append([str(k1) + str(k2)] + code)

    wb.save(f'{student}_{task_code}_{group}.xlsx')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    # Наименьшая вероятность
    min_p = 0.05
    # Наибольшая вероятность
    max_p = 0.95
    
    fn = 'list_magister_titpi_2020.txt'
    task_code = '07'

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if s:
            s_translit = pytils.translit.translify(s)
            print(s_translit)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                generator(group, student, task_code)
