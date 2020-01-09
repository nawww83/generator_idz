# Модуль генерирует файл с Заданием № 03
from openpyxl import Workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
from random import choice
from pprint import pprint as pp

mjr = sys.version_info.major
mnr = sys.version_info.minor
if (mjr == 3 and mnr < 7) or mjr < 3:
    print('Требуется Python версии 3.7 и выше!')
    exit()

student = 'IvanovAA'
task_code = '03'
group = '1B6'

# Ограничения на параметры (n, k) кода
min_n = 6
max_n = 15
min_k = 3
max_k = 5
min_r = 2

assert(min_k < min_n)
assert(max_k < max_n)

hf = Font(name = 'Calibri', bold = True)

# Генерация случайных (n, k) так, что k < n
while True:
    n = lc.randint(min_n, max_n)
    k = lc.randint(min_k, max_k)
    if k < n and (n - k) >= min_r:
        break

r = n - k

print(f'Введите нижнюю границу кодового расстояния (n, k)-кода ({n}, {k})')
print(f'Рекомендуется не более {max( (n - k + 1) // 2, 2 )}')
try:
    d_low_bound = int(input())
except:
    d_low_bound = 2
print(f'Подождите идет подбор порождающей матрицы G с кодовым \
расстоянием не ниже {d_low_bound}...')
G, d = lc.gen_matrix(n, k, d_low_bound)
Gsh, *_ = lc.shuffle_matrix(G, n, True, [])

print('Порождающая матрица G в систематической форме')
pp(G)

print('Матрица G после тасовки')
pp(Gsh)

a = lc.get_rand_bits(k)
s = lc.mult_v(a, Gsh)

print('Кодовое расстояние dк')
pp(d)

print('Выбранный кодовый вектор s')
pp(s)

qi = (d - 1) // 2 # Целевая кратность ошибки - кратность исправления
p = 1. * qi / n # Средняя кратность случайной величины q = np
e = lc.get_error_vector(n, p)
q = lc.hamming_weight(e) # Получившаяся кратность ошибки

print('Выпавший вектор ошибки e')
pp(e)

print('Кратность ошибки q')
pp(q)

v = lc.xor(s, e)

wb = Workbook()
ws = wb.active
ws.title = 'Main'

ws['A1'].font = hf
ws['A1'] = f'Порождающая матрица G (n, k)-кода ({n}, {k})'
for g_r in Gsh:
    ws.append(g_r)

ws.append(['Принятый кодовый вектор v'])
ws.append(v)

wsC = wb.create_sheet('Check')
wsC.append(['Введите ответы:'])
wsC.cell(row = wsC.max_row, column = 1).font = hf
wsC.append(['Проверочная матрица H:'])
for _ in range(r):
    wsC.append(lc.get_rand_bits(n))
wsC.append(['Кодовое расстояние кода dк:'])
wsC.append([0])

wsV = wb.create_sheet('CodeVector')
wsV.append(['Введите ответы:'])
wsV.cell(row = wsV.max_row, column = 1).font = hf
wsV.append(['Декодированный кодовый вектор s:'])
wsV.append(lc.get_rand_bits(n))
wsV.append(['Информационный вектор a:'])
wsV.append(lc.get_rand_bits(k))

wb.save(f'{student}_{task_code}_{group}.xlsx')
