# Модуль генерирует файл с Заданием № 05 на тему "Вероятность ошибки при 
# оптимальном приеме цифрового сигнала"
from openpyxl import Workbook
from openpyxl.styles import Font
import sys
from random import choice
from random import random
from pprint import pprint as pp
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

def generator(group, student, task_code):
    # Вид модуляции
    m = round( random() * (m_max - m_min) + m_min )

    print(f'Вид модуляции: {m}')

    # Длина кода
    n = round( random() * (n_max - n_min) + n_min )

    print(f'Длина кода: {n}')

    # Вероятность ошибки
    p = random() * (p_max - p_min) + p_min

    print(f'Вероятность битовой ошибки P бит. дек.: {p}')

    # Кратность исправления
    q = round( random() * (q_max - q_min) + q_min )

    print(f'Кратность исправления: {q}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'

    hf = Font(name = 'Calibri', bold = True)

    ws.append(['Вид модуляции m:'])
    ws.append([m])
    ws.append(['Длина кода n:'])
    ws.append([n])
    ws.append(['Требуемая вероятность битовой ошибки P бит. дек.:'])
    ws.append([p])
    ws.append(['Кратность исправления qи:'])
    ws.append([q])

    wsC = wb.create_sheet('Check')
    wsC.append(['Введите ответы:'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf
    wsC.append(['Скорость кодирования R:'])
    wsC.append([0.])
    wsC.append(['Отношение сигнал-шум на один бит, Eb/N0, дБ'])
    wsC.append([0.])
    wsC.append(['Внимание! Ответы давать с точностью не хуже 1%'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf

    wb.save(f'{student}_{task_code}_{group}.xlsx')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    fn = 'list_bakalavr_ots_2020.txt'
    task_code = '05'

    # Ограничения на вид модуляции
    m_min = 0
    m_max = 8

    modulations = {0: 'АМн когер.', 1: 'АМн некогер.', 2: 'ЧМн когер.', \
    3: 'ЧМн некогер.', 4: 'ФМн когер.', 5: 'ФМн част. когер.', \
    6: 'ФМн-4/КАМ-4 когер.', 7: 'ФМн-4/КАМ-4 част. когер', 8: 'ФМн-8 когер.'}

    # Ограничения на параметры (n, k) кода
    n_min = 31
    n_max = 255
    # Ограничение на вероятность битовой ошибки на выходе декодера
    p_min = 1.e-09
    p_max = 1.e-02

    # Ограничение на кратность исправления
    q_min = 1
    q_max = 7

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        s = s.replace('#', '')
        if s:
            s_translit = pytils.translit.translify(s)
            print(s_translit)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                generator(group, student, task_code)
