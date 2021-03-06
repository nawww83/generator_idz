# Модуль для проверки ответов к Заданию № 04
from openpyxl import load_workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
import operator
from functools import reduce
from pprint import pprint as pp
from copy import deepcopy
from scipy.special import comb
import numpy as np
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

# Проверяет состоит ли вектор только из 0 и 1
def is_bits_vector(v):
    return all(map(lambda x: (x == 0) or (x == 1), v))

# Проверяет состоит ли вектор только из чисел типа int
def is_int_vector(v):
    return all(map(lambda x: isinstance(x, int), v))

def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return

    ws = wb['Main']
    params = []
    for row in ws.iter_rows(min_row = 1, max_col = 1, max_row = 8, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row == 1 and (isinstance(row[0], int) or isinstance(row[0], float)):
            params.append(row[0])

    n_, k_, p_ = params
    assert(k_ < n_)
    assert(p_ <= p_max)
    assert(p_ >= p_min)

    r_ = n_ - k_

    wsC = wb['Check']
    params = []
    for row in wsC.iter_rows(min_row = 1, max_col = 1, max_row = 8, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row == 1 and (isinstance(row[0], int) or isinstance(row[0], float)):
            params.append(row[0])

    qi_, p_err_, p_bit_ = params

    assert(qi_ > 0)
    assert(qi_ <= r_ // 2)

    print(f'Параметры (n, k)-кода: ({n_}, {k_})')
    print(f'Число проверочных битов: {r_}')

    qi = lc.resolve_hamming_constrain(n_, k_)

    print(f'Введенная кратность исправляемой ошибки: qi = {qi_}')
    print(f'Правильный ответ: qi = {qi}')

    assert(qi_ == qi)

    p_err, _ = lc.probability_bsc_more(qi_, n_, p_)

    assert(p_err > 0.)

    print(f'Введенная вероятность ошибки: p_err = {p_err_}')
    print(f'Правильный ответ: p_err = {p_err}')

    rel_error = np.abs(p_err - p_err_) / p_err
    assert(rel_error < 0.01)

    d = 2 * qi + 1
    p_bit = p_err * d / n_

    print(f'Введенная вероятность битовой ошибки: p_bit = {p_bit_}')
    print(f'Правильный ответ: p_bit = {p_bit}')

    rel_error = np.abs(p_bit - p_bit_) / p_bit
    assert(rel_error < 0.01)

    print('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '04'
    fn = 'list_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 8
    max_n = 31
    min_k = 6
    max_k = 20
    min_r = 5
    # Ограничение на вероятность ошибки в BSC-канале
    p_min = 1.e-4
    p_max = 0.2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)
