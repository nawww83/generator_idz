# Модуль для проверки ответов к Заданию № 07
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
import error_rate as er
import source_codes as sc
import linear_codes as lc
from pprint import pprint as pp
import numpy as np
from random import random
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return

    wb = load_workbook(fname)

    ws = wb['Main']
    head_rows = 1 # Число строк на заголовок
    trash_rows = 1 # Число "мусорных" строк
    params = []
    for row in ws.iter_rows(min_row = 1, max_col = m_alphabet + 1, \
            max_row = trash_rows + 2 + 1, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row > 0:
            params.append(row)

    params = params[head_rows: ]
    sym_, P_ = params

    assert(sc.all_in_range_incl(P_, min_p, max_p))

    al_ = dict(zip(sym_, P_))
    print(f'Алфавит, {{X, P}}: {al_}')
    print(f'Норма: {sum(al_.values())}')


    wsC = wb['Check']
    head_rows = 3 # Число строк на заголовок
    trash_rows = 5 # Число "мусорных" строк
    params = []
    len_code_limit = 2 * int(np.log2(1. / min(P_)) + 0.5) + 1
    for row in wsC.iter_rows(min_row = 1, max_col = len_code_limit + 1, \
            max_row = trash_rows + 1 + m_alphabet + m_alphabet * m_alphabet, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row > 0:
            params.append(row)

    params = params[head_rows: ]

    symbols = []
    code_words = []
    for i in range(m_alphabet):
        symbols.append(params[i].pop(0))
        code_words.append(params[i])

    code_ = dict(zip(symbols, code_words))

    print('Введенный код Хаффмана для символа X:')
    print(code_)

    code = sc.make_huffman_table(al_)

    print('Найденный код Хаффмана (один из множества вариантов):')
    print(code)

    # Проверка правильности введенного ответа
    # По равенству длин кодовых слов
    l_ = {}
    for k, v in code_.items():
        l_[k] = len(v)
    l = {}
    for k, v in code.items():
        l[k] = len(v)

    assert(l_ == l)

    # По свойству префикса
    ok = sc.check_prefix(code_)

    assert(ok)

    params = params[-m_alphabet * m_alphabet: ]
    symbols = []
    code_words = []
    for i in range(m_alphabet * m_alphabet):
        symbols.append(params[i].pop(0))
        code_words.append(params[i])

    code_ = dict(zip(symbols, code_words))

    print('Введенный код Хаффмана для символа XX:')
    print(code_)

    alal = {}
    sym_match = {}
    for k1, v1 in al_.items():
        for k2, v2 in al_.items():
            sym_match[k1 + k2 * m_alphabet] = str(k2) + str(k1)
            alal[k1 + k2 * m_alphabet] = v1 * v2

    code_xx = sc.make_huffman_table(alal)
    code_xx = dict(sorted(code_xx.items()))

    code = {}
    for k, v in code_xx.items():
        code[sym_match[k]] = v

    print('Найденный код Хаффмана (один из множества вариантов):')
    print(code)

    # Проверка правильности введенного ответа
    # По равенству длин кодовых слов
    l_ = {}
    for k, v in code_.items():
        l_[k] = len(v)
    l = {}
    for k, v in code.items():
        l[k] = len(v)

    assert(l_ == l)

    # По свойству префикса
    ok = sc.check_prefix(code_)

    assert(ok)

    print('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '07'
    fn = 'list_magister_titpi_2020.txt'

    # Объем алфавита X
    m_alphabet = 3
    # Наименьшая вероятность
    min_p = 0.05
    # Наибольшая вероятность
    max_p = 0.95

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)