# Модуль для проверки ответов к Заданию № 02
from openpyxl import load_workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
import operator
from functools import reduce
from pprint import pprint as pp
from copy import deepcopy
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

# Проверяет состоит ли вектор только из 0 и 1
def is_bits_vector(v):
    return all(map(lambda x: (x == 0) or (x == 1), v))

# Проверяет состоит ли вектор только из чисел типа int
def is_int_vector(v):
    return all(map(lambda x: isinstance(x, int), v))


def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return

    ws = wb['Main']
    G = []
    for row in ws.iter_rows(min_row = 1, max_col = max_n, max_row = 2 + max_k, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        g_Ok = is_bits_vector(row)
        nr = len(row)
        if g_Ok and nr >= min_n and nr <= max_n:
            G.append(row) # Читаем матрицу кода

    k, n, ok = lc.check_matrix(G)
    r = n - k
    assert(ok)

    print(f'Порождающая матрица ({n}, {k})-кода')
    pp(G)

    parameters = []
    H_ = []
    wsC = wb['Check']
    for row in wsC.iter_rows(min_row = 1, max_col = max_n, max_row = 5 + max_n - min_k, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        h_Ok = is_bits_vector(row)
        nr = len(row)
        if h_Ok and nr >= min_n and nr <= max_n:
            H_.append(row) # Читаем матрицу кода
        elif nr == 1 and isinstance(row[0], int):
            parameters.append(row[0])

    r_, n_, ok = lc.check_matrix(H_)
    k_ = n_ - r_
    assert(r_ == r)
    assert(n_ == n)

    assert(ok)

    assert(sum(map(sum, H_)) > 0)

    assert((len(parameters) == 1))

    d_ = parameters[0]
    print(f'Идет вычисление кодового расстояния d...', flush = True)
    Wsp = lc.gen_spectrum(G)
    d = lc.spectrum_to_code_distance(Wsp)
    print(f'd = {d}', flush = True)
    print(f'Идет подбор проверочной матрицы H...', flush = True)
    H = lc.get_check_matrix(G)
    print(f'Идет вычисление кодового расстояния d по H...', flush = True)
    d_alter = lc.get_code_distance(H, False)

    print(f'Правильные ответы:')
    print(f'd = {d}, альтернативный метод d = {d_alter}')

    assert(d == d_alter)

    print(f'Введенные ответы:')
    pp(f'd = {d_}')

    assert(d_ == d)

    print(f'Правильная проверочная матрица кода')
    pp(H)

    print(f'Введенная проверочная матрица кода')
    pp(H_)

    zero = sum(map(sum, lc.mult_M(G, lc.transpose(H))))
    zero_ = sum(map(sum, lc.mult_M(G, lc.transpose(H_))))

    print(f'Контроль правильности проверочной матрицы')
    ok = (zero == zero_ == 0)
    pp(ok)
    assert(ok)

    pp('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '02'
    fn = 'list_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 6
    max_n = 15
    min_k = 3
    max_k = 5
    min_r = 2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)
