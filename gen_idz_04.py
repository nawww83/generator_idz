# Модуль генерирует файл с Заданием № 04 на тему "Неравенство Хемминга"
from openpyxl import Workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
from random import choice
from random import random
from pprint import pprint as pp
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

def generator(group, student, task_code):
    # Генерация случайных (n, k) так, что k < n
    while True:
        n = lc.randint(min_n, max_n)
        k = lc.randint(min_k, max_k)
        if k < n and (n - k) >= min_r:
            break
    r = n - k

    print(f'Параметры (n, k)-кода: ({n}, {k})')
    print(f'Число проверочных символов: {r}')

    # Генерация вероятности ошибки в канале
    while True:
        p = random()
        if p >= p_min and p <= p_max:
            break

    print(f'Вероятность ошибки в канале: {p}')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'

    hf = Font(name = 'Calibri', bold = True)

    ws['A1'].font = hf
    ws['A1'] = f'Параметры (n, k)-кода'
    ws.append(['Длина кода n'])
    ws.append([n])
    ws.append(['Количество информационных битов k'])
    ws.append([k])
    ws.append(['Вероятность ошибки в канале p'])
    ws.append([p])

    wsC = wb.create_sheet('Check')
    wsC.append(['Введите ответы:'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf
    wsC.append(['Кратность исправления qi:'])
    wsC.append([0])
    wsC.append(['Вероятность ошибки на выходе декодера P ош. дек.:'])
    wsC.append([0])
    wsC.append(['Вероятность битовой ошибки на выходе декодера P бит. дек.:'])
    wsC.append([0])
    wsC.append(['Внимание! Ответы давать с точностью не хуже 1%'])
    wsC.cell(row = wsC.max_row, column = 1).font = hf

    wb.save(f'{student}_{task_code}_{group}.xlsx')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '04'
    fn = 'list_bakalavr_ots_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 8
    max_n = 31
    min_k = 6
    max_k = 20
    min_r = 5

    # Ограничение на вероятность ошибки в BSC-канале
    p_min = 1.e-4
    p_max = 0.2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        s = s.replace('#', '')
        if s:
            s_translit = pytils.translit.translify(s)
            print(s_translit)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                generator(group, student, task_code)
