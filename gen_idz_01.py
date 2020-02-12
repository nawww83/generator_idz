# Модуль генерирует файл с Заданием № 01
from openpyxl import Workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
from pprint import pprint as pp
import numpy as np
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

def generator(group, student, task_code):
    # Генерация случайных (n, k) так, что k < n
    while True:
        n = lc.randint(min_n, max_n)
        k = lc.randint(min_k, max_k)
        if k < n and (n - k) >= min_r:
            break

    # print(f'Введите нижнюю границу кодового расстояния (n, k)-кода ({n}, {k})')
    d_recomend = lc.get_recomend_code_distance(n, k)
    # print(f'Рекомендуется не более {d_recomend}')
    #try:
    #    d_low_bound = int(input())
    #except:
    #    d_low_bound = 2
    print(f'Подождите идет подбор порождающей матрицы G с кодовым \
    расстоянием не ниже {d_recomend}...')
    G, _ = lc.gen_matrix(n, k, d_recomend)
    Gsh, *_ = lc.shuffle_matrix(G, n, True, [])

    print('Порождающая матрица G в систематической форме')
    pp(G)

    print('Матрица G после тасовки')
    pp(Gsh)

    Wsp = lc.gen_spectrum(G)
    print('Спектр кода Wsp по исходной матрице G (систематической)')
    pp(Wsp)

    Wspsh = lc.gen_spectrum(Gsh)
    print('Спектр кода Wsp по тасованной матрице G')
    pp(Wspsh)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Main'

    hf = Font(name = 'Calibri', bold = True)

    ws['A1'].font = hf
    ws['A1'] = 'Порождающая матрица G'
    for g_r in Gsh:
        ws.append(g_r)

    ws.append(['Введите ответы:'])
    ws.cell(row = ws.max_row, column = 1).font = hf
    ws.append(['Длина кода n:'])
    ws.append([0])
    ws.append(['Число информационных символов k:'])
    ws.append([0])
    ws.append(['Число проверочных символов r:'])
    ws.append([0])
    ws.append(['Кодовое расстояние кода dк:'])
    ws.append([0])
    ws.append(['Кратность гарантированного обнаружения qо:'])
    ws.append([0])
    ws.append(['Кратность гарантированного исправления qи:'])
    ws.append([0])

    spectr_name = 'CodeSpectrum'

    ws.append([f'Не забудьте заполнить лист {spectr_name}!'])
    ws.cell(row = ws.max_row, column = 1).font = hf

    wsWC = wb.create_sheet(spectr_name)
    wsWC['A1'].font = hf
    wsWC['A1'] = 'Введите спектр Wsp(C) кода (в отсортированном порядке)'
    wsWC.append(['Веса w:'])
    wsWC.append([i for i in range(n + 1)])
    wsWC.append(['Количество весов Nw:'])
    wsWC.append([0] * (n + 1))

    wb.save(f'{student}_{task_code}_{group}.xlsx')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '01'
    fn = 'list_magister_titpi_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 6
    max_n = 15
    min_k = 3
    max_k = 5
    min_r = 2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if s:
            s_translit = pytils.translit.translify(s)
            print(s_translit)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                generator(group, student, task_code)