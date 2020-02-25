# Модуль для проверки ответов к Заданию № 05
from openpyxl import load_workbook
from openpyxl.styles import Font
import sys
import error_rate as er
import linear_codes as lc
from pprint import pprint as pp
import numpy as np
import scipy.optimize as opt
from random import random
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return


    wb = load_workbook(fname)

    ws = wb['Main']
    params = []
    for row in ws.iter_rows(min_row = 1, max_col = 1, max_row = 9, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row == 1 and (isinstance(row[0], int) or isinstance(row[0], float)):
            params.append(row[0])

    m_, n_, p_, q_ = params
    assert(m_ >= m_min)
    assert(m_ <= m_max)
    assert(n_ >= n_min)
    assert(n_ <= n_max)
    assert(p_ <= p_max)
    assert(p_ >= p_min)
    assert(q_ <= q_max)
    assert(q_ >= q_min)

    print(f'Требуемая вероятность битовой ошибки P бит. дек.: {p_}')
    print(f'Длина кода n: {n_}')
    print(f'Кратность исправления: {q_}')
    print(f'Вид модуляции m: {m_}: {modulations[m_]}')

    wsC = wb['Check']
    params = []
    for row in wsC.iter_rows(min_row = 1, max_col = 1, max_row = 6, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        n_row = len(row)
        if n_row == 1 and isinstance(row[0], float) or isinstance(row[0], int):
            params.append(row[0])

    R_, EbN0_dB_ = params

    assert(R_ > 0.)
    assert(R_ <= 1.)

    print(f'Введенная скорость кодирования R: {R_}')
    print(f'Введенное отношение сигнал-шум на один бит Eb/N0: {EbN0_dB_}, дБ')

    EbN0_ = er.dB2pow(EbN0_dB_)
    M_ = bits_per_symbol[m_]
    EsN0_ = M_ * EbN0_ * R_

    p_sym_err, p_bit_err = error_func(m_)(EsN0_, coherencies[m_])

    print(f'Решение задачи оптимизации...')
    p_bit_dec = p_ # Целевая вероятность
    n = n_
    m = m_
    M = bits_per_symbol[m]

    R0 = 1. / n
    R1 = (n - 1.) / n
    RS = []
    for ir in range(472):
        rate = (R1 - R0) * ir / 471. + R0
        k = int(round(rate * n))
        assert(k > 0)
        assert(k < n)
        qi = lc.resolve_hamming_constrain(n, k)
        if qi != q_:
            continue
        d = 2 * qi + 1
        # Решение задачи оптимизации методом SciPy minimize
        def fun(x, args):
            p_target, m, qi, n = args
            _, pb = error_func(m)(x, coherencies[m])
            pe, _ = lc.probability_bsc_more(qi, n, pb)
            pb_dec = pe * d / n
            return np.fabs(np.log((pb_dec + 1.e-14) / (p_target + 1.e-14)))

        def f(x):
            return fun(x, (p_bit_dec, m, qi, n))

        rel_error = 1.
        min_EsN0 = 0.1
        max_EsN0 = 20.1
        while rel_error >= 0.01:
            init_EsN0 = random() * (max_EsN0 - min_EsN0) + min_EsN0
            EsN0 = opt.minimize(f, init_EsN0, \
                bounds = [(0.01, np.inf)], tol = 0.01).x[0]
            _, p_bit_err = error_func(m)(EsN0, coherencies[m])
            p_err, _ = lc.probability_bsc_more(qi, n, p_bit_err)
            p_bit = p_err * d / n
            rel_error = np.fabs(p_bit - p_bit_dec) / p_bit_dec
        RS.append((er.EbN0(EsN0, rate, M), rate, rel_error))
        #print(rate, er.EbN0(EsN0, rate, M), flush = True)

    EbN0_min = RS[0][0]
    R = RS[0][1]
    rel_error = RS[0][2]
    for rate in RS:
        if rate[0] < EbN0_min:
            EbN0_min = rate[0]
            R = rate[1]
            rel_error = rate[2]

    print(f'Целевая вероятность достигнута с относительной ошибкой: {rel_error:.3}')
    assert(rel_error < 0.01)

    k = int(round(R * n))
    qi = lc.resolve_hamming_constrain(n, k)

    print(f'Скорость кодирования R: {R}')
    EbN0_dB = er.pow2dB(EbN0_min)
    print(f'EbN0: {EbN0_dB:.4f} дБ')
    #
    print(f'P бит. дек.: {p_bit}')
    print(f'P бит. демод.: {p_bit_err}')
    print(f'P ош. дек.: {p_err}')
    print(f'Код (n, k): ({n}, {k})')
    print(f'Кратность исправления: {qi}')

    assert(qi == q_)

    rel_error = np.fabs(EbN0_dB - EbN0_dB_) / EbN0_dB
    assert(rel_error < 0.01)
    
    rel_error = np.fabs(R - R_) / R
    assert(rel_error < 0.01)

    print('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '05'
    fn = 'list_magister_titpi_2020.txt'

    # Ограничения на вид модуляции
    m_min = 0
    m_max = 8

    modulations = {0: 'АМн когер.', 1: 'АМн некогер.', 2: 'ЧМн когер.', \
    3: 'ЧМн некогер.', 4: 'ФМн когер.', 5: 'ФМн част. когер.', \
    6: 'ФМн-4/КАМ-4 когер.', 7: 'ФМн-4/КАМ-4 част. когер', 8: 'ФМн-8 когер.'}

    def error_func(i):
        switcher = {
            0: er.err_ask,
            1: er.err_ask,
            2: er.err_fsk,
            3: er.err_fsk,
            4: er.err_bpsk,
            5: er.err_bpsk,
            6: er.err_qpsk,
            7: er.err_qpsk,
            8: er.err_psk_8
        }
        return switcher.get(i, '')

    coherencies = {0: 0, 1: 1, 2: 0, 3: 1, 4: 0, 5: 1, 6: 0, 7: 1, 8: 0}

    # Число бит на символ
    bits_per_symbol = {0: 1, 1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 2, 7: 2, 8: 3}

    # Ограничения на параметры (n, k) кода
    n_min = 31
    n_max = 255
    # Ограничение на вероятность битовой ошибки на выходе декодера
    p_min = 1.e-09
    p_max = 1.e-02

    # Ограничение на кратность исправления
    q_min = 1
    q_max = 7

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)