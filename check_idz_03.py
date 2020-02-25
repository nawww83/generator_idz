# Модуль для проверки ответов к Заданию № 03
from openpyxl import load_workbook
from openpyxl.styles import Font
import linear_codes as lc
import sys
import operator
from functools import reduce
from pprint import pprint as pp
from copy import deepcopy
import pytils.translit
import re

def has_numbers(s):
    return re.search('\d', s)

# Проверяет состоит ли вектор только из 0 и 1
def is_bits_vector(v):
    return all(map(lambda x: (x == 0) or (x == 1), v))

# Проверяет состоит ли вектор только из чисел типа int
def is_int_vector(v):
    return all(map(lambda x: isinstance(x, int), v))

def checker(group, student, task_code):
    fname = f'{student}_{task_code}_{group}.xlsx'

    print(f'\n')
    print(f'*********************')
    print(f'Чтение файла {fname}')

    try:
        wb = load_workbook(fname)
    except FileNotFoundError:
        print(f'Файл {fname} не найден')
        return

    ws = wb['Main']
    G = []
    for row in ws.iter_rows(min_row = 1, max_col = max_n, max_row = 4 + max_k, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        g_Ok = is_bits_vector(row)
        n_row = len(row)
        if g_Ok and n_row >= min_n and n_row <= max_n:
            G.append(row) # Читаем матрицу кода

    # В конце считанной матрицы принятый кодовый вектор
    v = G.pop()

    print('Принятый кодовый вектор')
    pp(v)

    k, n, ok = lc.check_matrix(G)
    r = n - k
    assert(ok)

    print('Порождающая матрица кода')
    pp(G)

    parameters = []
    H_ = []
    wsC = wb['Check']
    for row in wsC.iter_rows(min_row = 1, max_col = max_n, max_row = 5 + max_n - min_k, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        h_Ok = is_bits_vector(row)
        n_row = len(row)
        if h_Ok and n_row >= min_n and n_row <= max_n:
            H_.append(row) # Читаем матрицу кода
        elif n_row == 1 and isinstance(row[0], int):
            parameters.append(row[0])

    r, n, ok = lc.check_matrix(H_)
    k = n - r
    assert(ok)

    assert((len(parameters) == 1))

    d_ = parameters[0]
    Wsp = lc.gen_spectrum(G)
    d = lc.spectrum_to_code_distance(Wsp)
    print('Подождите идет подбор матрицы H...')
    H = lc.get_check_matrix(G)
    d_alter = lc.get_code_distance(H, False)

    print('Правильные ответы:')
    pp(f'd = {d}, альтернативный метод d = {d_alter}')

    assert(d == d_alter)

    print('Введенные ответы:')
    pp(f'd = {d_}')

    assert(d_ == d)

    print('Правильная проверочная матрица кода')
    pp(H)

    print('Введенная проверочная матрица кода')
    pp(H_)

    zero = sum(map(sum, lc.mult_M(G, lc.transpose(H))))
    zero_ = sum(map(sum, lc.mult_M(G, lc.transpose(H_))))

    print('Контроль правильности проверочной матрицы')
    ok = (zero == zero_ == 0)
    pp(ok)
    assert(ok)

    wsV = wb['CodeVector']
    s_ = []
    a_ = []
    for row in wsV.iter_rows(min_row = 1, max_col = max_n, max_row = 6, values_only = True):
        row = list(filter(None.__ne__, row)) # Убирает ненужные None
        g_Ok = is_bits_vector(row)
        n_row = len(row)
        if g_Ok and n_row >= min_n and n_row <= max_n:
            s_.append(row) # Читаем декодированный кодовый вектор
        if g_Ok and n_row >= min_k and n_row <= max_k:
            a_.append(row) # Читаем декодированный информационный вектор

    assert(len(s_) == 1)
    assert(len(a_) == 1)

    s_ = s_[0]
    a_ = a_[0]

    n_ = len(s_)
    k_ = len(a_)

    assert(n_ == n)
    assert(k_ == k)

    ac = lc.get_min_adjacent_classes(H)
    c = lc.mult_v(v, lc.transpose(H))
    e = ac[tuple(c)]
    s_est = lc.xor(v, e)

    print('Скорректированный кодовый вектор')
    pp(s_est)

    print('Синдром')
    pp(c)

    print('Вектор ошибки минимальной кратности')
    pp(e)

    print('Введенный скорректированный кодовый вектор')
    pp(s_)

    zero = sum(lc.mult_v(s_est, lc.transpose(H)))
    zero_ = sum(lc.mult_v(s_, lc.transpose(H_)))

    assert(zero == zero_ == 0)

    e_ = lc.xor(v, s_)
    qe = lc.hamming_weight(e)
    qe_ = lc.hamming_weight(e_)

    assert(qe == qe_)

    print('Найденный вектор ошибки наименьшей кратности')
    pp(e)
    print('Кратность')
    pp(qe)

    print('Вектор ошибки, соответствующий введенному кодовому вектору')
    pp(e_)
    print('Кратность')
    pp(qe_)

    a_est = lc.decode(s_est, G)
    a__ = lc.decode(s_, G)

    print('Информационный вектор')
    pp(a__)
    pp(s_)

    print('Введенный информационный вектор')
    pp(a_)
    pp(s_)

    assert(a__ == a_)

    pp('All Ok')

if __name__ == "__main__":
    mjr = sys.version_info.major
    mnr = sys.version_info.minor
    if (mjr == 3 and mnr < 7) or mjr < 3:
        print('Требуется Python версии 3.7 и выше!')
        exit()

    task_code = '03'
    fn = 'list_magister_titpi_2020.txt'

    # Ограничения на параметры (n, k) кода
    min_n = 6
    max_n = 15
    min_k = 3
    max_k = 5
    min_r = 2

    assert(min_k < min_n)
    assert(max_k < max_n)

    students_file = open(fn, 'r', encoding = 'utf-8')
    students = students_file.readlines()
    group = ''
    student = ''
    for s in students:
        s = s.strip()
        if '#' in s:
            continue
        if s:
            s_translit = pytils.translit.translify(s)
            if has_numbers(s_translit):
                group = s_translit
            else:
                student = s_translit
                checker(group, student, task_code)